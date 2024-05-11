from rest_framework import status
from rest_framework.views import APIView
from .models import AktivitiesProject
from .models import EngineerActivity, StakeholderActivity, AktivitiesProject
from .serializers import EngineerActivitySerializer, StakeholderActivitySerializer
from rest_framework import generics, filters
from .models import AktivitiesProject, EngineerActivity, StakeholderActivity
from .models import AktivitiesProject, EngineerActivity, StakeholderActivity
from rest_framework import viewsets, status
from rest_framework.response import Response

from .serializers import AktivitiesProjectSerializer, ActivityDeleteSerializer


class AktivitiesProjectViewSet(generics.ListCreateAPIView):
    queryset = AktivitiesProject.objects.all()
    serializer_class = AktivitiesProjectSerializer

class EngineerActivityViewSet(generics.ListCreateAPIView):
    queryset = EngineerActivity.objects.all()
    serializer_class = EngineerActivitySerializer

# Tambah dan Buat
class StakeholderActivityViewSet(generics.ListCreateAPIView):
    queryset = StakeholderActivity.objects.all()
    serializer_class = StakeholderActivitySerializer

# Hapus
class UserDeleteView(generics.DestroyAPIView):
    queryset = AktivitiesProject.objects.all()
    serializer_class = ActivityDeleteSerializer

# Update
class AktivitiesProjectRetrieveUpdateView(generics.RetrieveUpdateAPIView):
    queryset = AktivitiesProject.objects.all()
    serializer_class = AktivitiesProjectSerializer

# Detail
class AktivitiesProjectDetailView(generics.RetrieveAPIView):
    queryset = AktivitiesProject.objects.all()
    serializer_class = AktivitiesProjectSerializer

#Search 
class ActivityListSearch(generics.ListCreateAPIView):
    queryset = AktivitiesProject.objects.all()
    serializer_class = AktivitiesProjectSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name'] 

# Filter Status
class AktivitiesProjectListView(APIView):
    def post(self, request):
        selected_status = request.data.get('selected_status', [])  # Mendapatkan daftar grup yang dipilih
        # Lakukan filter berdasarkan grup yang dipilih
        filtered_activity = AktivitiesProject.objects.filter(status__in=selected_status)
        # Serialize data pengguna yang difilter
        serialized_activity = AktivitiesProjectSerializer(filtered_activity, many=True, context={'request': request})
        return Response(serialized_activity.data, status=status.HTTP_200_OK)

# Eksport Excel
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from .models import AktivitiesProject, EngineerActivity, StakeholderActivity
from .serializers import AktivitiesProjectSerializer, EngineerActivitySerializer, StakeholderActivitySerializer

class ExportAktivitiesProjectExcel(APIView):
    def get(self, request):
        # Ambil data Aktivitas Proyek dari basis data
        aktivitas_projects = AktivitiesProject.objects.all()

        # Inisialisasi workbook dan worksheet
        wb = Workbook()
        ws = wb.active

        # Header untuk file Excel
        header = [
            'Aktivitas Proyek ID', 'Project', 'PM', 'Name', 'Date Start', 'Date Finish',
            'Date Estimated', 'Note', 'Status', 'Tanggung Jawab', 
            'Engineer', '', '',  # Engineer columns
            'Stakeholder', '', '',  # Stakeholder columns
        ]
        ws.append(header)

        # Header anak untuk baris kedua
        child_header = [
            '','','','','','','','','','','Nama Engineer', 'Status', 'Prsentase',  # Engineer
            'Nama User', 'Stakeholder',  # Stakeholder
        ]
        ws.append(child_header)

        # Menggabungkan sel 
        ws.merge_cells('K1:M1')  # Engineer
        ws.merge_cells('N1:O1')  # Stakeholder

        # Menambahkan data ke dalam file Excel
        for aktivitas_project in aktivitas_projects:
            # Hapus zona waktu dari objek datetime
            date_start = aktivitas_project.date_start.astimezone(None).strftime("%Y-%m-%d %H:%M:%S")
            date_finish = aktivitas_project.date_finish.astimezone(None).strftime("%Y-%m-%d %H:%M:%S")
            date_estimated = aktivitas_project.date_estimated.astimezone(None).strftime("%Y-%m-%d %H:%M:%S")

            # Tambahkan data Aktivitas Proyek
            ws.append([
                aktivitas_project.id,
                aktivitas_project.project.name,
                aktivitas_project.pm.username,
                aktivitas_project.name,
                date_start,
                date_finish,
                date_estimated,
                aktivitas_project.note,
                aktivitas_project.status,
                aktivitas_project.tanggung_jawab,
            ])

            # Ambil data Aktivitas Engineer dan tambahkan ke dalam baris yang sama
            engineer_activities = EngineerActivity.objects.filter(activity=aktivitas_project)
            for engineer_activity in engineer_activities:
                ws.append([
                    '',  # ID Proyek (kosong)
                    '',  # Project (kosong)
                    '',  # PM (kosong)
                    '',  # Name (kosong)
                    '',  # Date Start (kosong)
                    '',  # Date Finish (kosong)
                    '',  # Date Estimated (kosong)
                    '',  # Note (kosong)
                    '',  # Status (kosong)
                    '',  # Tanggung Jawab (kosong)
                    engineer_activity.engineer.username if engineer_activity.engineer else '',  # Engineer
                    engineer_activity.status if engineer_activity.status else '',  # Engineer Status
                    engineer_activity.persentase_beban_kerja if engineer_activity.persentase_beban_kerja else '',  # Engineer Persentase Beban
                    '',  # Stakeholder (kosong)
                    '',  # Stakeholder User (kosong)
                    '',  # Stakeholder Stakeholder (kosong)
                ])

            # Ambil data Aktivitas Stakeholder dan tambahkan ke dalam baris yang sama
            stakeholder_activities = StakeholderActivity.objects.filter(activity=aktivitas_project)
            for stakeholder_activity in stakeholder_activities:
                ws.append([
                    '',  # ID Proyek (kosong)
                    '',  # Project (kosong)
                    '',  # PM (kosong)
                    '',  # Name (kosong)
                    '',  # Date Start (kosong)
                    '',  # Date Finish (kosong)
                    '',  # Date Estimated (kosong)
                    '',  # Note (kosong)
                    '',  # Status (kosong)
                    '',  # Tanggung Jawab (kosong)
                    '',  # Engineer (kosong)
                    '',  # Engineer Status (kosong)
                    '',  # Engineer Persentase Beban (kosong)
                    stakeholder_activity.stakeholder if stakeholder_activity.stakeholder else '',  # Stakeholder
                    stakeholder_activity.user.username if stakeholder_activity.user else '',  # Stakeholder User
                    stakeholder_activity.stakeholder if stakeholder_activity.stakeholder else '',  # Stakeholder Stakeholder
                ])

        # Menyimpan workbook ke dalam HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=aktivitas_projects.xlsx'
        wb.save(response)
        return response








# @api_view(['GET'])
# def list_all_aktivities_project(request):
#     aktivities_projects = AktivitiesProject.objects.all()
#     serializer = AktivitasSerializer(aktivities_projects, many=True)
#     return Response(serializer.data)


# @api_view(['POST'])
# def create_aktivity_project(request):
#     serializer = AktivitasSerializer(data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data, status=status.HTTP_201_CREATED)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# @api_view(['PUT'])
# def update_aktivity_project(request, pk):
#     try:
#         aktivity_project = AktivitiesProject.objects.get(pk=pk)
#     except AktivitiesProject.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = AktivitasSerializer(aktivity_project, data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# @api_view(['DELETE'])
# def delete_aktivity_project(request, pk):
#     try:
#         aktivity_project = AktivitiesProject.objects.get(pk=pk)
#     except AktivitiesProject.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     aktivity_project.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)


# @api_view(['GET'])
# def list_engineers_for_activity(request, activity_id):
#     try:
#         engineers = EngineerActivity.objects.filter(activity_id=activity_id)
#         serializer = EngineerActivitySerializer(engineers, many=True)
#         return Response(serializer.data)
#     except EngineerActivity.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)
