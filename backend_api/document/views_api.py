from rest_framework import generics, filters
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import ProjectDocument
from .serializers import ProjectDocumentSerializer
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from django.http import HttpResponse
from .models import ProjectDocument
from django.utils import timezone
from rest_framework.parsers import MultiPartParser, FormParser 

from django.http import FileResponse
import tempfile

import csv
import json
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

# views_api.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from import_export.widgets import ForeignKeyWidget
from tablib import Dataset
from datetime import datetime

def export_documents_to_pdf(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Retrieve documents data
    documents = ProjectDocument.objects.all()

    # Filter documents by month if provided
    if month:
        documents = documents.filter(upload_date__month=month)

    # Filter documents by year if provided
    if year:
        documents = documents.filter(upload_date__year=year)

    # If there are no documents found, return a JSON response indicating so
    if not documents.exists():
        return JsonResponse({'message': 'No data available for the selected month and year.'}, status=404)

    # Create a buffer to store PDF data
    buffer = BytesIO()

    table_width = 500  # You should adjust this based on your actual table width
    left_margin = (letter[1] - table_width) / 2
    right_margin = (letter[1] - table_width) / 2

    # Create the PDF object, using the buffer as its "file"
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=left_margin, rightMargin=right_margin)

    # Set up PDF content
    styles = getSampleStyleSheet()
    style = ParagraphStyle('TableHeader', parent=styles['Heading2'], textColor=colors.black, spaceAfter=6, alignment=1)
    pdf_title = Paragraph("Documents Data", style)

    # Define column names manually
    field_names = ['Project', 'Uploader', 'Name', 'Upload Date', 'Category', 'Description', 'Document']

    # Set up data rows
    data = [field_names]
    for document in documents:
        row = [
            document.project.name if document.project else '',
            document.uploader.email if document.uploader else '',
            document.name,
            document.upload_date.strftime('%Y-%m-%d'),
            document.category if document.category else '',
            document.description if document.description else '',
            request.build_absolute_uri(document.document_file.url)
        ]
        data.append(row)

    # Create table
    table = Table(data, splitByRow=1)

    # Add style to the table
    style = TableStyle(
        [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Ensure vertical alignment to middle
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),  # Add inner grid for better visibility
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ]
    )
    table.setStyle(style)

    # Build the PDF
    elements = [pdf_title, table]
    pdf.build(elements)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file
    response = HttpResponse(content_type='application/pdf')
    if month and year:
        response['Content-Disposition'] = 'attachment; filename=documents_{0}_{1}.pdf'.format(month, year)
    elif month:
        response['Content-Disposition'] = 'attachment; filename=documents_{0}.pdf'.format(month)
    elif year:
        response['Content-Disposition'] = 'attachment; filename=documents_{0}.pdf'.format(year)
    else:
        response['Content-Disposition'] = 'attachment; filename=documents_all.pdf'

    response.write(buffer.getvalue())

    return response


def export_documents_to_json(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Retrieve documents data
    documents = ProjectDocument.objects.all()

    # Filter documents by month if provided
    if month:
        documents = documents.filter(upload_date__month=month)

    # Filter documents by year if provided
    if year:
        documents = documents.filter(upload_date__year=year)

    # If there are no documents found, return a JSON response indicating so
    if not documents.exists():
        return JsonResponse({'message': 'No data available for the selected month and year.'}, status=404)

    # Convert data to a list of dictionaries
    data_list = []
    for document in documents:
        data_list.append({
            'Id': document.id,
            'Project': document.project.name if document.project else '',
            'Uploader': document.uploader.email if document.uploader else '',
            'Name': document.name,
            'Upload Date': document.upload_date.strftime('%Y-%m-%d'),
            'Category': document.category if document.category else '',
            'Description': document.description if document.description else '',
            'Created At': document.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Updated At': document.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Download Link': request.build_absolute_uri(document.document_file.url),
        })

    # Convert data to JSON
    json_data = json.dumps(data_list, indent=2)

    # Set the filename based on filter parameters
    if month and year:
        filename = 'documents_{0}_{1}.json'.format(month, year)
    elif month:
        filename = 'documents_{0}.json'.format(month)
    elif year:
        filename = 'documents_{0}.json'.format(year)
    else:
        filename = 'documents_all.json'

    # Create the HttpResponse object with the appropriate JSON header
    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename={0}'.format(filename)

    return response


def export_documents_to_csv(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    if month and year:
        filename = 'documents_{0}_{1}.csv'.format(month, year)
    elif month:
        filename = 'documents_{0}.csv'.format(month)
    elif year:
        filename = 'documents_{0}.csv'.format(year)
    else:
        filename = 'documents_all.csv'
    response['Content-Disposition'] = 'attachment; filename={0}'.format(filename)

    # Create a CSV writer using the HttpResponse object as the "file."
    writer = csv.writer(response)

    # Write the header row to the CSV file
    field_names = ['Id', 'Project', 'Uploader', 'Name', 'Upload Date', 'Category', 'Description', 'Created At', 'Updated At', 'Download Link']
    writer.writerow(field_names[:-1])  # Exclude the 'Download Link' field

    # Retrieve documents data
    documents = ProjectDocument.objects.all()

    # Filter documents by month if provided
    if month:
        documents = documents.filter(upload_date__month=month)

    # Filter documents by year if provided
    if year:
        documents = documents.filter(upload_date__year=year)

    # If there are no documents found, return a JSON response indicating so
    if not documents.exists():
        return JsonResponse({'message': 'No data available for the selected month and year.'}, status=404)

    # Write data rows to the CSV file
    for document in documents:
        # Convert upload_date to string without timezone information
        upload_date_str = document.upload_date.strftime('%Y-%m-%d')

        # Create a list with formatted values
        row_data = [
            document.id,
            document.project.name if document.project else '',
            document.uploader.email if document.uploader else '',
            document.name,
            upload_date_str,
            document.category if document.category else '',
            document.description if document.description else '',
            document.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            document.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            request.build_absolute_uri(document.document_file.url),  # Download Link
        ]

        # Write the row to the CSV file
        writer.writerow(row_data[:-1])  # Exclude the 'Download Link' field

    return response



def export_documents_to_excel(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column names manually
    field_names = ['Id', 'Project', 'Uploader', 'Name', 'Upload Date', 'Category', 'Description', 'Created At', 'Updated At', 'Download Link']

    # Write headers to the worksheet
    ws.append(field_names[:-1])  # Exclude the 'Download Link' field

    # Retrieve documents data
    documents = ProjectDocument.objects.all()

    # Filter documents by month if provided
    if month:
        documents = documents.filter(upload_date__month=month)

    # Filter documents by year if provided
    if year:
        documents = documents.filter(upload_date__year=year)

    # If there are no documents found, return a JSON response indicating so
    if not documents.exists():
        return JsonResponse({'message': 'No data available for the selected month and year.'}, status=404)

    # Create a named style for hyperlinks
    hyperlink_style = NamedStyle(name='hyperlink_style', font=Font(color="0000FF", underline='single'))

    # Write data to the worksheet
    for document, row_num in zip(documents, range(2, len(documents) + 2)):
        # Convert upload_date to string without timezone information
        upload_date_str = document.upload_date.strftime('%Y-%m-%d')

        # Create a list with formatted values
        row_data = [
            document.id,
            document.project.name if document.project else '',
            document.uploader.email if document.uploader else '',
            document.name,
            upload_date_str,
            document.category if document.category else '',
            document.description if document.description else '',
            document.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            document.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            request.build_absolute_uri(document.document_file.url),  # Download Link
        ]

        for col_num, value in enumerate(row_data):
            ws.cell(row=row_num, column=col_num + 1).value = value

        # Add a hyperlink to the 'Download Link' column
        download_link_cell = ws.cell(row=row_num, column=len(field_names)).value
        ws.cell(row=row_num, column=len(field_names)).hyperlink = download_link_cell
        ws.cell(row=row_num, column=len(field_names)).style = hyperlink_style

    # Save workbook to a temporary file
    tmp_file = tempfile.NamedTemporaryFile(delete=False)
    wb.save(tmp_file.name)
    tmp_file.close()

    # Serve the file using Django FileResponse
    response = FileResponse(open(tmp_file.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if month and year:
        response['Content-Disposition'] = 'attachment; filename=documents_{0}_{1}.xlsx'.format(month, year)
    elif month:
        response['Content-Disposition'] = 'attachment; filename=documents_{0}.xlsx'.format(month)
    elif year:
        response['Content-Disposition'] = 'attachment; filename=documents_{0}.xlsx'.format(year)
    else:
        response['Content-Disposition'] = 'attachment; filename=documents_all.xlsx'

    return response

# Detail doc
class ProjectDocumentDetail(generics.RetrieveAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

#Create dan List
class ProjectDocumentListCreate(generics.ListCreateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

# Edit
class ProjectDocumentRetrieveUpdate(generics.RetrieveUpdateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

    def put(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

# Delete
class ProjectDocumentDestroy(generics.DestroyAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

# Search
class ProjectDocumentListSearch(generics.ListCreateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['category', 'name'] 

# ASC, DSC
class ProjectDocumentList(generics.ListCreateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

    def list(self, request, *args, **kwargs):
        # Mendapatkan nilai 'order_by' dan 'order' dari parameter query string
        order_by = request.query_params.get('order_by', 'name')
        order = request.query_params.get('order', 'asc')

        # Validasi nilai 'order'
        if order not in ['asc', 'desc']:
            return Response({"error": "Invalid order value. Use 'asc' or 'desc'."}, status=400)

        # Validasi nilai 'order_by'
        valid_order_by_fields = ['name', 'project_id', 'uploader_id' , 'upload_date', 'category']  
        if order_by not in valid_order_by_fields:
            return Response({"error": f"Invalid order_by value. Use one of {valid_order_by_fields}."}, status=400)

        # Mengurutkan queryset berdasarkan bidang yang dipilih
        queryset = self.filter_queryset(self.get_queryset().order_by(f"{'-' if order == 'desc' else ''}{order_by}"))

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    
# Filter berdasarkan bulan dan/atau tahun
class DocumentListFilterByMonthYear(generics.ListAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')

        if month and year:
            try:
                month = int(month)
                year = int(year)
                queryset = queryset.filter(upload_date__month=month, upload_date__year=year)
            except ValueError:
                return JsonResponse({"error": "Invalid month or year value."}, status=400)
        elif month:
            try:
                month = int(month)
                queryset = queryset.filter(upload_date__month=month)
            except ValueError:
                return JsonResponse({"error": "Invalid month value."}, status=400)
        elif year:
            try:
                year = int(year)
                queryset = queryset.filter(upload_date__year=year)
            except ValueError:
                return JsonResponse({"error": "Invalid year value."}, status=400)

        return queryset

# Filter berdasarkan category
class DocumentListFilterByCategory(generics.ListAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        category = self.request.query_params.get('category')

        if category:
            queryset = queryset.filter(category=category)

        return queryset

# Count data
class DocumentCount(APIView):
    def get(self, request):
        document_count = ProjectDocument.objects.count()
        return Response({'count': document_count})