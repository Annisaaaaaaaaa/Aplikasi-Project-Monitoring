from rest_framework import generics, filters
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Invoice
from .serializers import InvoiceSerializer

from django.http import HttpResponse
from openpyxl import Workbook

from openpyxl.styles import NamedStyle, Font
from django.http import FileResponse
import tempfile


import csv
import json
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

# views_api.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from import_export.widgets import ForeignKeyWidget
from tablib import Dataset
from .resources import InvoiceResource
from datetime import datetime


# Import API endpoint
@csrf_exempt
def import_invoices(request):
    if request.method == 'POST':
        dataset = Dataset()

        # Set headers for the dataset
        dataset.headers = ['project', 'to_contact', 'sent_date', 'due_date', 'date', 'amount', 'status', 'note', 'document_file']

        new_invoices = request.FILES['file']

        if not new_invoices.name.endswith('xlsx'):
            return JsonResponse({'error': 'File must be in Excel (xlsx) format.'})

        imported_data = dataset.load(new_invoices.read(), 'xlsx')

        for data_row in imported_data:
            # ... (lanjutkan debug di bagian lain)

            # Convert date strings to datetime objects if needed
            date_fields = ['sent_date', 'due_date', 'date']
            for field in date_fields:
                if field in data_row and data_row[field]:
                    try:
                        data_row[field] = datetime.strptime(data_row[field], '%Y-%m-%d').date()
                    except ValueError:
                        return JsonResponse({'error': f'Invalid date format for {field}.'})

            # Creating a resource instance
            resource = InvoiceResource()
            
            # Pass the dataset directly to import_data
            result = resource.import_data(dataset, dry_run=True, raise_errors=False)

            if not result.has_errors():
                # You can choose to use dry_run=False if the import is successful
                # resource.import_data(dataset, dry_run=False)

                return JsonResponse({'success': 'Invoices imported successfully.'})
            else:
                return JsonResponse({'error': 'There was an error importing the file.'})

    return JsonResponse({'error': 'Invalid request method.'})


def export_invoices_to_pdf(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Retrieve invoices data
    invoices = Invoice.objects.all()

    # Filter invoices by month if provided
    if month:
        invoices = invoices.filter(date__month=month)

    # Filter invoices by year if provided
    if year:
        invoices = invoices.filter(date__year=year)

    # Create a buffer to store PDF data
    buffer = BytesIO()

    table_width = 500  # You should adjust this based on your actual table width
    left_margin = (letter[1] - table_width) / 2
    right_margin = (letter[1] - table_width) / 2

    # Create the PDF object
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=left_margin, rightMargin=right_margin)

    # Set up PDF content
    styles = getSampleStyleSheet()
    style = ParagraphStyle('TableHeader', parent=styles['Heading2'], textColor=colors.black, spaceAfter=6, alignment=1)
    pdf_title = Paragraph("Invoices Data", style)

    # Define column names manually
    field_names = ['Project', 'To Contact', 'Sent Date', 'Due Date', 'Date', 'Amount', 'Status', 'Note', 'Download Link']

    # Set up data rows
    data = [field_names]
    for invoice in invoices:
        row = [
            invoice.project.name if invoice.project else '',
            invoice.to_contact.name if invoice.to_contact else '',
            str(invoice.sent_date) if invoice.sent_date else '',
            str(invoice.due_date) if invoice.due_date else '',
            str(invoice.date),
            str(invoice.amount) if invoice.amount else '',
            invoice.get_status_display(),
            invoice.note if invoice.note else '',
            request.build_absolute_uri(invoice.document_file.url)
        ]
        data.append(row)

    # Create table
    table = Table(data, repeatRows=1)

    # Add style to the table
    style = TableStyle(
        [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ]
    )
    table.setStyle(style)

    # Build the PDF
    elements = [pdf_title, table]
    pdf.build(elements)

    # FileResponse sets the Content-Disposition header for file download
    response = HttpResponse(content_type='application/pdf')
    if month:
        response['Content-Disposition'] = 'attachment; filename=invoices_{0}.pdf'.format(month)
    elif year:
        response['Content-Disposition'] = 'attachment; filename=invoices_{0}.pdf'.format(year)
    else:
        response['Content-Disposition'] = 'attachment; filename=invoices_all.pdf'

    response.write(buffer.getvalue())

    return response


def export_invoices_to_json(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Retrieve invoices data
    invoices = Invoice.objects.all()

    # Filter invoices by month if provided
    if month:
        invoices = invoices.filter(date__month=month)

    # Filter invoices by year if provided
    if year:
        invoices = invoices.filter(date__year=year)

    # Convert data to a list of dictionaries
    data_list = []
    for invoice in invoices:
        data_list.append({
            'Project': invoice.project.name if invoice.project else '',
            'To Contact': invoice.to_contact.name if invoice.to_contact else '',
            'Sent Date': str(invoice.sent_date) if invoice.sent_date else '',
            'Due Date': str(invoice.due_date) if invoice.due_date else '',
            'Date': str(invoice.date),
            'Amount': str(invoice.amount) if invoice.amount else '',
            'Status': invoice.get_status_display(),
            'Note': invoice.note if invoice.note else '',
            'Download Link': request.build_absolute_uri(invoice.document_file.url),
        })

    # Convert data to JSON
    json_data = json.dumps(data_list, indent=2)

    # Create the HttpResponse object with the appropriate JSON header
    response = HttpResponse(json_data, content_type='application/json')
    if month:
        response['Content-Disposition'] = 'attachment; filename=invoices_{0}.json'.format(month)
    elif year:
        response['Content-Disposition'] = 'attachment; filename=invoices_{0}.json'.format(year)
    else:
        response['Content-Disposition'] = 'attachment; filename=invoices_all.json'

    return response



def export_invoices_to_csv(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')

    # Set the filename based on the filter, if any
    if month:
        filename = f'invoices_{month}.csv'
    elif year:
        filename = f'invoices_{year}.csv'
    else:
        filename = 'invoices_all.csv'

    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Create a CSV writer using the HttpResponse object as the "file."
    writer = csv.writer(response)

    # Write the header row to the CSV file
    field_names = ['Project', 'To Contact', 'Sent Date', 'Due Date', 'Date', 'Amount', 'Status', 'Note', 'Download Link']
    writer.writerow(field_names[:-1])  # Exclude the 'Download Link' field

    # Retrieve invoices data
    invoices = Invoice.objects.all()

    # Filter invoices by month if provided
    if month:
        invoices = invoices.filter(date__month=month)

    # Filter invoices by year if provided
    if year:
        invoices = invoices.filter(date__year=year)

    # Write data rows to the CSV file
    for invoice in invoices:
        # Convert date fields to string without timezone information
        sent_date_str = invoice.sent_date.strftime('%Y-%m-%d') if invoice.sent_date else ''
        due_date_str = invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else ''
        date_str = invoice.date.strftime('%Y-%m-%d') if invoice.date else ''

        # Create a list with formatted values
        row_data = [
            invoice.project.name if invoice.project else '',
            invoice.to_contact.name if invoice.to_contact else '',
            sent_date_str,
            due_date_str,
            date_str,
            str(invoice.amount) if invoice.amount else '',
            invoice.get_status_display(),
            invoice.note if invoice.note else '',
            request.build_absolute_uri(invoice.document_file.url),  # Download Link
        ]

        # Write the row to the CSV file
        writer.writerow(row_data[:-1])  # Exclude the 'Download Link' field

    return response



def export_invoices_to_excel(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column names manually
    field_names = ['Project', 'To Contact', 'Sent Date', 'Due Date', 'Date', 'Amount', 'Status', 'Note', 'Download Link']

    # Write headers to the worksheet
    ws.append(field_names[:-1])

    # Retrieve invoices data
    invoices = Invoice.objects.all()

    # Filter invoices by month if provided
    if month:
        invoices = invoices.filter(date__month=month)

    # Filter invoices by year if provided
    if year:
        invoices = invoices.filter(date__year=year)

    # Create a named style for hyperlinks
    hyperlink_style = NamedStyle(name='hyperlink_style', font=Font(color="0000FF", underline='single'))

    # Write data to the worksheet
    for invoice, row_num in zip(invoices, range(2, len(invoices) + 2)):
        # Convert date fields to string without timezone information
        sent_date_str = invoice.sent_date.strftime('%Y-%m-%d') if invoice.sent_date else ''
        due_date_str = invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else ''
        date_str = invoice.date.strftime('%Y-%m-%d') if invoice.date else ''

        # Create a list with formatted values
        row_data = [
            invoice.project.name if invoice.project else '',
            invoice.to_contact.name if invoice.to_contact else '',
            sent_date_str,
            due_date_str,
            date_str,
            str(invoice.amount) if invoice.amount else '',
            invoice.get_status_display(),
            invoice.note if invoice.note else '',
            request.build_absolute_uri(invoice.document_file.url),  
        ]

        for col_num, value in enumerate(row_data):
            ws.cell(row=row_num, column=col_num + 1).value = value

        # Add a hyperlink to the 'Download Link' column
        download_link_cell = ws.cell(row=row_num, column=len(field_names)).value
        ws.cell(row=row_num, column=len(field_names)).hyperlink = download_link_cell
        ws.cell(row=row_num, column=len(field_names)).style = hyperlink_style

    # Save workbook to a temporary file
    tmp_file = tempfile.NamedTemporaryFile(delete=False)
    wb.save(tmp_file.name)
    tmp_file.close()

    # Serve the file using Django FileResponse
    response = FileResponse(open(tmp_file.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if month:
        response['Content-Disposition'] = f'attachment; filename=invoices_{month}.xlsx'
    elif year:
        response['Content-Disposition'] = f'attachment; filename=invoices_{year}.xlsx'
    else:
        response['Content-Disposition'] = 'attachment; filename=invoices_all.xlsx'

    return response

# Detail invoice
class InvoiceDetail(generics.RetrieveAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

#Create dan List
class InvoiceListCreate(generics.ListCreateAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

# Edit
class InvoiceRetrieveUpdate(generics.RetrieveUpdateAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

# Delete
class InvoiceDestroy(generics.DestroyAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

# Search
class InvoiceListSearch(generics.ListCreateAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'project'] 


# ASC, DSC
class InvoiceList(generics.ListCreateAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

    def list(self, request, *args, **kwargs):
        # Mendapatkan nilai 'order_by' dan 'order' dari parameter query string
        order_by = request.query_params.get('order_by', 'status')
        order = request.query_params.get('order', 'asc')

        # Validasi nilai 'order'
        if order not in ['asc', 'desc']:
            return Response({"error": "Invalid order value. Use 'asc' or 'desc'."}, status=400)

        # Validasi nilai 'order_by'
        valid_order_by_fields = ['status', 'project_id', 'due_date']  
        if order_by not in valid_order_by_fields:
            return Response({"error": f"Invalid order_by value. Use one of {valid_order_by_fields}."}, status=400)

        # Mengurutkan queryset berdasarkan bidang yang dipilih
        queryset = self.filter_queryset(self.get_queryset().order_by(f"{'-' if order == 'desc' else ''}{order_by}"))

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# Filter berdasarkan bulan dan/atau tahun
class InvoiceListFilterByMonthYear(generics.ListAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')

        if month and year:
            try:
                month = int(month)
                year = int(year)
                queryset = queryset.filter(date__month=month, date__year=year)
            except ValueError:
                return JsonResponse({"error": "Invalid month or year value."}, status=400)
        elif month:
            try:
                month = int(month)
                queryset = queryset.filter(date__month=month)
            except ValueError:
                return JsonResponse({"error": "Invalid month value."}, status=400)
        elif year:
            try:
                year = int(year)
                queryset = queryset.filter(date__year=year)
            except ValueError:
                return JsonResponse({"error": "Invalid year value."}, status=400)

        return queryset

# Filter berdasarkan status
class InvoiceListFilterByStatus(generics.ListAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        status = self.request.query_params.get('status')

        if status:
            queryset = queryset.filter(status=status)

        return queryset
    
# Filter berdasarkan type
class InvoiceListFilterByType(generics.ListAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        type = self.request.query_params.get('type')

        if type:
            queryset = queryset.filter(type=type)

        return queryset

# Count data
class InvoiceCount(APIView):
    def get(self, request):
        invoice_count = Invoice.objects.count()
        return Response({'count': invoice_count})