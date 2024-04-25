from rest_framework import generics, filters
from rest_framework.response import Response
from .models import Client
from .serializers import ClientSerializer

from django.http import HttpResponse
from openpyxl import Workbook

from openpyxl.styles import NamedStyle, Font
from django.http import FileResponse
import tempfile

import csv
import json
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from import_export.widgets import ForeignKeyWidget
from tablib import Dataset
from .resources import ClientResource
from datetime import datetime

#Create dan List
class ClientListCreate(generics.ListCreateAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

# Edit
from rest_framework import status

from rest_framework import status
from rest_framework.response import Response

class ClientRetrieveUpdate(generics.RetrieveUpdateAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        # Jika 'logo' tidak ada dalam data permintaan atau logo tidak berubah,
        # hapus 'logo' dari data yang divalidasi agar logo yang ada tetap dipertahankan
        if 'logo' not in request.data or request.data['logo'] == instance.logo:
            serializer.validated_data.pop('logo', None)

        # Memperbarui data objek dengan data yang diterima
        self.perform_update(serializer)

        # Mengembalikan serializer.data yang telah diperbarui
        return Response(serializer.data)


# Delete
class ClientDestroy(generics.DestroyAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

# Search
class ClientListSearch(generics.ListCreateAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name'] 

# ASC, DSC
class ClientList(generics.ListCreateAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

    def list(self, request, *args, **kwargs):
        # Mendapatkan nilai 'order_by' dan 'order' dari parameter query string
        order_by = request.query_params.get('order_by', 'name')
        order = request.query_params.get('order', 'asc')

        # Validasi nilai 'order'
        if order not in ['asc', 'desc']:
            return Response({"error": "Invalid order value. Use 'asc' or 'desc'."}, status=400)

        # Validasi nilai 'order_by'
        valid_order_by_fields = ['name', 'industry', 'company_size' , 'status', 'date_joined']  
        if order_by not in valid_order_by_fields:
            return Response({"error": f"Invalid order_by value. Use one of {valid_order_by_fields}."}, status=400)

        # Mengurutkan queryset berdasarkan bidang yang dipilih
        queryset = self.filter_queryset(self.get_queryset().order_by(f"{'-' if order == 'desc' else ''}{order_by}"))

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    # Import API endpoint
@csrf_exempt
def import_clients(request):
    if request.method == 'POST':
        dataset = Dataset()

        # Set headers for the dataset
        dataset.headers = ['name', 'pic_phone', 'pic_email', 'pic_title',
            'industry', 'website_url', 'logo', 'company_size', 'company_address',
            'company_email', 'company_phone', 'additional_info',
            'date_joined', 'status']

        new_projects = request.FILES['file']

        if not new_projects.name.endswith('xlsx'):
            return JsonResponse({'error': 'File must be in Excel (xlsx) format.'})

        imported_data = dataset.load(new_projects.read(), 'xlsx')

        for data_row in imported_data:
            # ... (lanjutkan debug di bagian lain)

            # Convert date strings to datetime objects if needed
            date_fields = ['date_joined']
            for field in date_fields:
                if field in data_row and data_row[field]:
                    try:
                        data_row[field] = datetime.strptime(data_row[field], '%Y-%m-%d').date()
                    except ValueError:
                        return JsonResponse({'error': f'Invalid date format for {field}.'})

            # Creating a resource instance
            resource = ClientResource()
            
            # Pass the dataset directly to import_data
            result = resource.import_data(dataset, dry_run=True, raise_errors=False)

            if not result.has_errors():
                # You can choose to use dry_run=False if the import is successful
                # resource.import_data(dataset, dry_run=False)

                return JsonResponse({'success': 'clients imported successfully.'})
            else:
                return JsonResponse({'error': 'There was an error importing the file.'})

    return JsonResponse({'error': 'Invalid request method.'})

from django.http import HttpResponse, JsonResponse, FileResponse
from django.utils.html import escape
from io import BytesIO
from reportlab.lib.pagesizes import landscape, letter, inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
import csv
import json
from .models import Client

def export_clients_to_csv(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=clients.csv'

    # Create a CSV writer using the HttpResponse object as the "file."
    writer = csv.writer(response)

    # Write the header row to the CSV file
    field_names = ['Name', 'PIC Phone', 'PIC Email', 'PIC Title',
                   'Industry', 'Website URL', 'Company Size', 'Company Address',
                   'Company Email', 'Company Phone', 'Additional Info',
                   'Date Joined', 'Status', 'Logo']
    writer.writerow(field_names)

    # Write data rows to the CSV file
    clients = Client.objects.all()

    for client in clients:
        date_joined_str = client.date_joined.strftime('%Y-%m-%d') if client.date_joined else ''
        logo_url = request.build_absolute_uri(client.logo.url) if client.logo else ''

        # Create a list with formatted values
        row_data = [
            client.name if client.name else '',
            client.pic_phone if client.pic_phone else '',
            client.pic_email if client.pic_email else '',
            client.pic_title if client.pic_title else '',
            client.industry if client.industry else '',
            client.website_url if client.website_url else '',
            client.company_size if client.company_size else '',
            client.company_address if client.company_address else '',
            client.company_email if client.company_email else '',
            client.company_phone if client.company_phone else '',
            client.additional_info if client.additional_info else '',
            date_joined_str,
            client.get_status_display(),
            logo_url,
        ]

        # Write the row to the CSV file
        writer.writerow(row_data)

    return response

def export_clients_to_json(request):
    # Retrieve clients data
    clients = Client.objects.all()

    # Convert data to a list of dictionaries
    data_list = []
    for client in clients:
        data_list.append({
            'Name': client.name if client.name else '',
            'PIC Phone': client.pic_phone if client.pic_phone else '',
            'PIC Email': client.pic_email if client.pic_email else '',
            'PIC Title': client.pic_title if client.pic_title else '',
            'Industry': client.industry if client.industry else '',
            'Website URL': client.website_url if client.website_url else '',
            'Company Size': client.company_size if client.company_size else '',
            'Company Address': client.company_address if client.company_address else '',
            'Company Email': client.company_email if client.company_email else '',
            'Company Phone': client.company_phone if client.company_phone else '',
            'Additiona Info': client.additional_info if client.additional_info else '',
            'Date Joined': str(client.date_joined) if client.date_joined else '',
            'Status': client.get_status_display(),
            'Logo': request.build_absolute_uri(client.logo.url) if client.logo else '',
        })

    # Convert data to JSON
    json_data = json.dumps(data_list, indent=2)

    # Create the HttpResponse object with the appropriate JSON header
    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename=clients.json'

    return response

def export_clients_to_pdf(request):
    # Retrieve clients data
    clients = Client.objects.all()

    # Create a buffer to store PDF data
    buffer = BytesIO()

    # Create the PDF object
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter))

    # Set up PDF content
    styles = getSampleStyleSheet()
    style = ParagraphStyle('TableHeader', parent=styles['Heading2'], textColor=colors.black, spaceAfter=6, alignment=1)
    pdf_title = Paragraph("Clients Data", style)

    # Define column names manually
    field_names = [
        'Name', 'PIC Phone', 'PIC Email', 'PIC Title',
        'Industry', 'Website URL', 'Company Size', 'Company Address',
        'Company Email', 'Company Phone', 'Additional Info',
        'Date Joined', 'Status', 'Logo'
    ]

    # Set up data rows
    data = [field_names]
    for client in clients:
        date_joined_str = client.date_joined.strftime('%Y-%m-%d') if client.date_joined else ''
        logo_url = request.build_absolute_uri(client.logo.url) if client.logo else ''
        row = [
            client.name if client.name else '',
            client.pic_phone if client.pic_phone else '',
            client.pic_email if client.pic_email else '',
            client.pic_title if client.pic_title else '',
            client.industry if client.industry else '',
            client.website_url if client.website_url else '',
            client.company_size if client.company_size else '',
            client.company_address if client.company_address else '',
            client.company_email if client.company_email else '',
            client.company_phone if client.company_phone else '',
            client.additional_info if client.additional_info else '',
            date_joined_str,
            client.get_status_display(),
            escape(logo_url),
        ]
        data.append(row)

    # Create table
    table = Table(data, repeatRows=1)

    # Add style to the table
    style = TableStyle(
        [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ]
    )
    table.setStyle(style)

    # Build the PDF
    elements = [pdf_title, Spacer(1, 0.2*inch), table]  # Add spacer to create space between title and table
    pdf.build(elements)

    # FileResponse sets the Content-Disposition header for file download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=clients.pdf'
    response.write(buffer.getvalue())

    return response

def export_clients_to_excel(request):
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column names manually
    field_names = ['Name', 'PIC Phone', 'PIC Email', 'PIC Title',
                   'Industry', 'Website URL', 'Company Size', 'Company Address',
                   'Company Email', 'Company Phone', 'Additional Info',
                   'Date Joined', 'Status', 'Logo']

    # Write headers to the worksheet
    ws.append(field_names)

    clients = Client.objects.all()

    hyperlink_style = NamedStyle(name='hyperlink_style', font=Font(color="0000FF", underline='single'))

    # Write data to the worksheet
    for client, row_num in zip(clients, range(2, len(clients) + 2)):
        date_joined_str = client.date_joined.strftime('%Y-%m-%d') if client.date_joined else ''
        logo_url = request.build_absolute_uri(client.logo.url) if client.logo else ''

        # Create a list with formatted values
        row_data = [
            client.name if client.name else '',
            client.pic_phone if client.pic_phone else '',
            client.pic_email if client.pic_email else '',
            client.pic_title if client.pic_title else '',
            client.industry if client.industry else '',
            client.website_url if client.website_url else '',
            client.company_size if client.company_size else '',
            client.company_address if client.company_address else '',
            client.company_email if client.company_email else '',
            client.company_phone if client.company_phone else '',
            client.additional_info if client.additional_info else '',
            date_joined_str,
            client.get_status_display(),
            logo_url,
        ]

        for col_num, value in enumerate(row_data):
            if col_num == len(field_names) - 1:  # Check if it's the 'Logo' column
                # Add hyperlink as text
                ws.cell(row=row_num, column=col_num + 1).value = value
            else:
                # Add other values as usual
                ws.cell(row=row_num, column=col_num + 1).value = value

        # Set hyperlink style for the 'Logo' column
        ws.cell(row=row_num, column=len(field_names)).style = hyperlink_style

    # Add filter to header row
    ws.auto_filter.ref = ws.dimensions

    # Create a temporary file to save the workbook
    temp_file = BytesIO()
    wb.save(temp_file)

    # Create the response with the Excel file
    response = HttpResponse(
        temp_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=clients.xlsx'

    return response
