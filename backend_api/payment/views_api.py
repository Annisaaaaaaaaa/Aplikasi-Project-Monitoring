from rest_framework import generics, filters
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Payment
from .serializers import PaymentSerializer

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Payment

from openpyxl.styles import NamedStyle, Font
from django.http import FileResponse
import tempfile
from invoice.resources import InvoiceResource

import csv
import json
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

# views_api.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from import_export.widgets import ForeignKeyWidget
from tablib import Dataset
from datetime import datetime


import tempfile
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
from django.http import FileResponse
from datetime import datetime

# Import API endpoint
@csrf_exempt
def import_payments(request):
    if request.method == 'POST':
        dataset = Dataset()

        # Set headers for the dataset
        dataset.headers = ['project', 'to_contact', 'sent_date', 'due_date', 'date', 'amount', 'status', 'note', 'document_file']

        new_payments = request.FILES['file']

        if not new_payments.name.endswith('xlsx'):
            return JsonResponse({'error': 'File must be in Excel (xlsx) format.'})

        imported_data = dataset.load(new_payments.read(), 'xlsx')

        for data_row in imported_data:
            # ... (lanjutkan debug di bagian lain)

            # Convert date strings to datetime objects if needed
            date_fields = ['sent_date', 'due_date', 'date']
            for field in date_fields:
                if field in data_row and data_row[field]:
                    try:
                        data_row[field] = datetime.strptime(data_row[field], '%Y-%m-%d').date()
                    except ValueError:
                        return JsonResponse({'error': f'Invalid date format for {field}.'})

            # Creating a resource instance
            resource = InvoiceResource()
            
            # Pass the dataset directly to import_data
            result = resource.import_data(dataset, dry_run=True, raise_errors=False)

            if not result.has_errors():
                # You can choose to use dry_run=False if the import is successful
                # resource.import_data(dataset, dry_run=False)

                return JsonResponse({'success': 'Invoices imported successfully.'})
            else:
                return JsonResponse({'error': 'There was an error importing the file.'})

    return JsonResponse({'error': 'Invalid request method.'})


def export_payments_to_pdf(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Retrieve payments data
    payments = Payment.objects.all()

    # Filter payments by month if provided
    if month:
        payments = payments.filter(payment_date__month=month)

    # Filter payments by year if provided
    if year:
        payments = payments.filter(payment_date__year=year)

    # Create a buffer to store PDF data
    buffer = BytesIO()

    table_width = 500  # You should adjust this based on your actual table width
    left_margin = (letter[1] - table_width) / 2
    right_margin = (letter[1] - table_width) / 2

    # Create the PDF object
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=left_margin, rightMargin=right_margin)

    # Set up PDF content
    styles = getSampleStyleSheet()
    style = ParagraphStyle('TableHeader', parent=styles['Heading2'], textColor=colors.black, spaceAfter=6, alignment=1)
    pdf_title = Paragraph("Payment Data", style)

    # Define column names manually
    field_names = ['Project', 'Payer Name', 'Payer Account', 'Receiver Name', 'Receiver Account', 'Amount', 'Payment Date', 'Note', 'Download Link']

    # Set up data rows
    data = [field_names]
    for payment in payments:
        payment_date_str = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else ''
        row = [
            payment.project.name if payment.project else '',
            payment.payer_name if payment.payer_name else '',
            payment.payer_account_number if payment.payer_account_number else '',
            payment.receiver_name if payment.receiver_name else '',
            payment.receiver_account_number if payment.receiver_account_number else '',
            str(payment.amount) if payment.amount else '',
            payment_date_str,
            payment.note if payment.note else '',
            request.build_absolute_uri(payment.document_file.url)
        ]
        data.append(row)

    # Create table
    table = Table(data, repeatRows=1)

    # Add style to the table
    style = TableStyle(
        [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ]
    )
    table.setStyle(style)

    # Build the PDF
    elements = [pdf_title, table]
    pdf.build(elements)

    # FileResponse sets the Content-Disposition header for file download
    response = HttpResponse(content_type='application/pdf')
    if month and year:
        response['Content-Disposition'] = 'attachment; filename=payments_{0}_{1}.pdf'.format(month, year)
    elif month:
        response['Content-Disposition'] = 'attachment; filename=payments_{0}.pdf'.format(month)
    elif year:
        response['Content-Disposition'] = 'attachment; filename=payments_{0}.pdf'.format(year)
    else:
        response['Content-Disposition'] = 'attachment; filename=payments_all.pdf'

    response.write(buffer.getvalue())

    return response


from datetime import datetime

def export_payments_to_json(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Retrieve payments data
    payments = Payment.objects.all()

    # Filter payments by month if provided
    if month:
        payments = payments.filter(payment_date__month=month)

    # Filter payments by year if provided
    if year:
        payments = payments.filter(payment_date__year=year)

    # Convert data to a list of dictionaries
    data_list = []
    for payment in payments:
        payment_date_str = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else ''
        data_list.append({
            'Project': payment.project.name if payment.project else '',
            'Payer Name': payment.payer_name if payment.payer_name else '',
            'Payer Account': payment.payer_account_number if payment.payer_account_number else '',
            'Receiver Name': payment.receiver_name if payment.receiver_name else '',
            'Receiver Account': payment.receiver_account_number if payment.receiver_account_number else '',
            'Amount': str(payment.amount) if payment.amount else '',
            'Date': payment_date_str,
            'Note': payment.note if payment.note else '',
            'File Link': request.build_absolute_uri(payment.document_file.url),
        })

    # Convert data to JSON
    json_data = json.dumps(data_list, indent=2)

    # Determine the filename based on filter parameters
    filename = 'payments.json'
    if month and year:
        filename = f'payments_{year}_{month}.json'
    elif month:
        filename = f'payments_{month}.json'
    elif year:
        filename = f'payments_{year}.json'

    # Create the HttpResponse object with the appropriate JSON header
    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response


import csv
from datetime import datetime

def export_payments_to_csv(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    filename = 'payments.csv'

    # Adjust filename based on filter parameters
    if month and year:
        filename = f'payments_{year}_{month}.csv'
    elif month:
        filename = f'payments_{month}.csv'
    elif year:
        filename = f'payments_{year}.csv'

    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Create a CSV writer using the HttpResponse object as the "file."
    writer = csv.writer(response)

    # Write the header row to the CSV file
    field_names = ['Project', 'Payment Date', 'Amount', 'Note', 'Payer Name', 'Payer Account Number', 'Receiver Name', 'Receiver Account Number', 'Download Link']
    writer.writerow(field_names[:-1])  # Exclude the 'Download Link' field

    # Retrieve payments data based on filter parameters
    payments = Payment.objects.all()

    if month:
        payments = payments.filter(payment_date__month=month)

    if year:
        payments = payments.filter(payment_date__year=year)

    # Write data rows to the CSV file
    for payment in payments:
        # Convert payment_date to string without timezone information
        payment_date_str = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else ''

        # Create a list with formatted values
        row_data = [
            payment.project.name if payment.project else '',
            payment_date_str,
            str(payment.amount) if payment.amount else '',
            payment.note if payment.note else '',
            payment.payer_name if payment.payer_name else '',
            payment.payer_account_number if payment.payer_account_number else '',
            payment.receiver_name if payment.receiver_name else '',
            payment.receiver_account_number if payment.receiver_account_number else '',
            request.build_absolute_uri(payment.document_file.url) if payment.document_file else '',
        ]

        # Write the row to the CSV file
        writer.writerow(row_data[:-1])  # Exclude the 'Download Link' field

    return response



def export_payments_to_excel(request):
    # Retrieve filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column names manually
    field_names = ['Project', 'Payment Date', 'Amount', 'Note', 'Payer Name', 'Payer Account Number', 'Receiver Name', 'Receiver Account Number', 'Download Link']

    # Write headers to the worksheet
    ws.append(field_names[:-1] + ['document_file'])  

    # Retrieve payments data based on filter parameters
    payments = Payment.objects.all()

    if month:
        payments = payments.filter(payment_date__month=month)

    if year:
        payments = payments.filter(payment_date__year=year)

    # Create a named style for hyperlinks
    hyperlink_style = NamedStyle(name='hyperlink_style', font=Font(color="0000FF", underline='single'))

    # Write data to the worksheet
    for payment, row_num in zip(payments, range(2, len(payments) + 2)):
        # Convert payment_date to string without timezone information
        payment_date_str = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else ''

        # Create a list with formatted values
        row_data = [
            payment.project.name if payment.project else '',
            payment_date_str,
            str(payment.amount) if payment.amount else '',
            payment.note if payment.note else '',
            payment.payer_name if payment.payer_name else '',
            payment.payer_account_number if payment.payer_account_number else '',
            payment.receiver_name if payment.receiver_name else '',
            payment.receiver_account_number if payment.receiver_account_number else '',
            request.build_absolute_uri(payment.document_file.url) if payment.document_file else '',
        ]

        for col_num, value in enumerate(row_data):
            ws.cell(row=row_num, column=col_num + 1).value = value

        # Add a hyperlink to the 'Download Link' column
        download_link_cell = ws.cell(row=row_num, column=len(field_names)).value
        ws.cell(row=row_num, column=len(field_names)).hyperlink = download_link_cell
        ws.cell(row=row_num, column=len(field_names)).style = hyperlink_style

    # Save workbook to a temporary file
    tmp_file = tempfile.NamedTemporaryFile(delete=False)
    wb.save(tmp_file.name)
    tmp_file.close()

    # Serve the file using Django FileResponse
    response = FileResponse(open(tmp_file.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'payments.xlsx'

    # Adjust filename based on filter parameters
    if month and year:
        filename = f'payments_{year}_{month}.xlsx'
    elif month:
        filename = f'payments_{month}.xlsx'
    elif year:
        filename = f'payments_{year}.xlsx'

    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response

# Detail payment
class PaymentDetail(generics.RetrieveAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

#Create dan List
class PaymentListCreate(generics.ListCreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

# Edit
class PaymentRetrieveUpdate(generics.RetrieveUpdateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer


# Delete
class PaymentDestroy(generics.DestroyAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer


# Search
class PaymentListSearch(generics.ListCreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'payer_name'] 

# ASC, DSC
class PaymentList(generics.ListCreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

    def list(self, request, *args, **kwargs):
        # Mendapatkan nilai 'order_by' dan 'order' dari parameter query string
        order_by = request.query_params.get('order_by', 'payment_date')
        order = request.query_params.get('order', 'asc')

        # Validasi nilai 'order'
        if order not in ['asc', 'desc']:
            return Response({"error": "Invalid order value. Use 'asc' or 'desc'."}, status=400)

        # Validasi nilai 'order_by'
        valid_order_by_fields = ['payment_date', 'payer_name', 'receiver_name']  
        if order_by not in valid_order_by_fields:
            return Response({"error": f"Invalid order_by value. Use one of {valid_order_by_fields}."}, status=400)

        # Mengurutkan queryset berdasarkan bidang yang dipilih
        queryset = self.filter_queryset(self.get_queryset().order_by(f"{'-' if order == 'desc' else ''}{order_by}"))

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    

# Filter berdasarkan bulan dan/atau tahun
class PaymentListFilterByMonthYear(generics.ListAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')

        if month and year:
            try:
                month = int(month)
                year = int(year)
                queryset = queryset.filter(payment_date__month=month, date__year=year)
            except ValueError:
                return JsonResponse({"error": "Invalid month or year value."}, status=400)
        elif month:
            try:
                month = int(month)
                queryset = queryset.filter(payment_date__month=month)
            except ValueError:
                return JsonResponse({"error": "Invalid month value."}, status=400)
        elif year:
            try:
                year = int(year)
                queryset = queryset.filter(payment_date__year=year)
            except ValueError:
                return JsonResponse({"error": "Invalid year value."}, status=400)

        return queryset

# Filter berdasarkan type
class PaymentListFilterByType(generics.ListAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        type = self.request.query_params.get('type')

        if type:
            queryset = queryset.filter(type=type)

        return queryset

# Count data
class PaymentCount(APIView):
    def get(self, request):
        payment_count = Payment.objects.count()
        return Response({'count': payment_count})